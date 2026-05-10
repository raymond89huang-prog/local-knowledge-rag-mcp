import csv
from pathlib import Path
from typing import Iterable, List

from doc_rag.document import DocumentChunk
from .base import DocumentLoader


class SpreadsheetLoader(DocumentLoader):
    supported_extensions = {".csv", ".xlsx"}

    def load(self, file_path: Path) -> List[DocumentChunk]:
        if file_path.suffix.lower() == ".csv":
            return self._load_csv(file_path)
        return self._load_xlsx(file_path)

    def _load_csv(self, file_path: Path) -> List[DocumentChunk]:
        with file_path.open("r", encoding="utf-8-sig", errors="ignore", newline="") as f:
            rows = list(csv.reader(f))
        return self._rows_to_chunks(file_path, file_path.stem, "csv", rows)

    def _load_xlsx(self, file_path: Path) -> List[DocumentChunk]:
        from openpyxl import load_workbook

        workbook = load_workbook(str(file_path), read_only=True, data_only=True)
        chunks: List[DocumentChunk] = []
        for sheet in workbook.worksheets:
            rows = [[self._format_cell(value) for value in row] for row in sheet.iter_rows(values_only=True)]
            chunks.extend(self._rows_to_chunks(file_path, sheet.title, "xlsx", rows))
        workbook.close()
        return chunks

    def _rows_to_chunks(
        self,
        file_path: Path,
        sheet_name: str,
        file_type: str,
        rows: List[List[str]],
        batch_size: int = 40,
    ) -> List[DocumentChunk]:
        chunks: List[DocumentChunk] = []
        title = file_path.stem
        rows = [row for row in rows if any(str(cell).strip() for cell in row)]
        if not rows:
            return chunks

        header = rows[0]
        data_rows = rows[1:] if len(rows) > 1 else rows
        for start in range(0, len(data_rows), batch_size):
            batch = data_rows[start : start + batch_size]
            row_start = start + 2 if len(rows) > 1 else start + 1
            row_end = row_start + len(batch) - 1
            content = self._format_table(header, batch)
            if len(content) <= 10:
                continue
            location = f"{sheet_name} rows {row_start}-{row_end}"
            chunks.append(
                DocumentChunk(
                    content=content,
                    heading=location,
                    path=str(file_path),
                    title=title,
                    metadata={
                        "file_type": file_type,
                        "sheet": sheet_name,
                        "row_start": row_start,
                        "row_end": row_end,
                        "location": location,
                    },
                )
            )
        return chunks

    @staticmethod
    def _format_table(header: Iterable[str], rows: List[List[str]]) -> str:
        header_values = [str(cell).strip() for cell in header]
        lines = [" | ".join(header_values)]
        for row in rows:
            values = [str(cell).strip() for cell in row]
            lines.append(" | ".join(values))
        return "\n".join(lines)

    @staticmethod
    def _format_cell(value) -> str:
        if value is None:
            return ""
        return str(value)

